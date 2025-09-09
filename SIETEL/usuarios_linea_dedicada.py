#!/usr/bin/env python
# coding: utf-8

# In[43]:


import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns


# Listar el número de las hojas

# In[2]:


archivo_excel = pd.ExcelFile('entrada/USUARIOS.xlsx')
numero_hojas = len(archivo_excel.sheet_names)
numero_hojas


# Listar el nombre de las hojas

# In[3]:


nombre_hojas = archivo_excel.sheet_names
nombre_hojas


# In[4]:


for hojas in nombre_hojas:
    print (hojas)


# # Lectura de Archivo Abril

# <b>Mes de evaluación</b>

# In[179]:


nombre_hojas[0]


# In[180]:


prestador_abril = pd.read_excel('entrada/USUARIOS.xlsx', sheet_name = nombre_hojas[0])


# In[181]:


prestador_abril.head()


# ## Dimensiones

# In[182]:


prestador_abril.shape


# In[183]:


prestador_abril.info()


# In[184]:


prestador_abril.columns.values


# In[185]:


prestador_abril.isna().sum()


# In[186]:


print(f'Mes analizado: {nombre_hojas[0]}')


# In[187]:


print (f'Número total de registros de usuarios es: {prestador_abril.shape[0]}')


# In[188]:


prestador_abril['tipo enlace'].value_counts().reset_index()


# ## 1. Cantidad Usuarios por Parroquia

# In[189]:


client_parroquia_abril = (
    prestador_abril.groupby(['parroquia'])['CLIENTE']
        .count()
        .reset_index()
        .sort_values(by='CLIENTE', ascending=False)
)
client_parroquia_abril


# In[190]:


client_parroquia_abril.columns


# <b>Frecuencia de usuarios por parroquia</b>

# In[308]:


frecuencia_user_parroquia_abril = (
    prestador_abril.groupby(['parroquia'])['CLIENTE']
        .value_counts()
        .reset_index()
)
frecuencia_user_parroquia_abril


# In[309]:


frecuencia_user_parroquia_abril.columns


# <b>Filtro por clientes</b>

# In[193]:


cliente = ['FABIAN RUIZ OLMEDO']
filtro = prestador_abril['CLIENTE'].isin(cliente)
prestador_abril[filtro]


# In[194]:


cliente = ['JAVIER LEONARDO CAMPOVERDE RAMIREZ']
filtro = prestador_abril['CLIENTE'].isin(cliente)
prestador_abril[filtro]


# In[195]:


cliente = ['JORDAN ALEXANDER MEJIA CRUZ']
filtro = prestador_abril['CLIENTE'].isin(cliente)
prestador_abril[filtro]


# In[196]:


cliente = ['JUAN BELTRAN ARBOLEDA']
filtro = prestador_abril['CLIENTE'].isin(cliente)
prestador_abril[filtro]


# In[197]:


cliente = ['LENIN CHANATASIG QUIMBITA']
filtro = prestador_abril['CLIENTE'].isin(cliente)
prestador_abril[filtro]


# Nota: Realizar comparación entre los clientes activos y estos usuarios

# Realizar comparación de la cantidad de usuarios a lo largo de los 3 meses

# ## 2. Clientes por parroquias

# Datos a graficar

# In[198]:


prestador_count = (
    prestador_abril.groupby(['parroquia'])['CLIENTE']
        .count()
        .reset_index()
        .sort_values(by='CLIENTE', ascending=False)
)


# In[200]:


# Configuración inicial
plt.figure(figsize=(10, 8))
sns.set_theme(style="whitegrid")

# Barras Horizontal
ax = sns.barplot(
    data=prestador_count,  # DataFrame con los datos agrupados
    y='parroquia',
    x='CLIENTE',
    hue = 'parroquia',
    legend = False,
    orient='h'
)

# Añadir etiquetas (valores de las barras)
for container in ax.containers:
    ax.bar_label(container, fmt='%.0f', padding=3)

# Personalización
plt.title('No. Clientes por Parroquia - Abril', fontsize=16, pad=20)
plt.xlabel('Número de Clientes', fontsize=12)
plt.ylabel('Parroquia', fontsize=12)

# Ajustar layout y mostrar
plt.tight_layout()
plt.show()



# Etiquetas para las barras

# Etiqueta posición [0]

# In[201]:


ax.bar_label(ax.containers[0], fmt='%.0f', padding=3)  # Agregar valores en las barras


# Etiqueta todas las posiciones

# In[202]:


for container in ax.containers:
    print(ax.bar_label(container, fmt='%.0f', padding=3))


# ## 3. Distribución de Velocidad de Internet (Megas)

# In[204]:


# Análisis de distribución de megas
plt.figure(figsize=(12, 6))
megas_dist = prestador_abril['megas'].value_counts().sort_index()

# Gráfico Barras Verticales
plt.subplot(1, 2, 1)
sns.barplot(x=megas_dist.index, y=megas_dist.values, palette='coolwarm')
plt.title('Distribución de Planes de Megas - Abril')
plt.xlabel('Megas')
plt.ylabel('Número de Clientes')
plt.xticks(rotation=45)

# Agregar etiquetas (valores de las barras)
for i, v in enumerate(megas_dist.values):
    plt.text(i, v + 5, str(v), ha='center')

# Gráfico Pie
plt.subplot(1, 2, 2)
plt.pie(megas_dist.values, labels=megas_dist.index, autopct='%1.1f%%')
plt.title('Porcentaje de Clientes por Plan - Abril')

# Ajustar Layout y mostrar
plt.tight_layout()
plt.show()


# ## 4. Análisis de Tipo de Enlace

# Distribución por tipo de enlace

# In[205]:


# Distribución por tipo de enlace
tipo_enlace_count = prestador_abril['tipo enlace'].value_counts().reset_index()
tipo_enlace_count.columns = ['tipo_enlace', 'count']


# In[207]:


plt.figure(figsize=(8, 4))
ax = sns.barplot(data=tipo_enlace_count, y='tipo_enlace', x='count', palette='Set2')
plt.title('Distribución por Tipo de Enlace - Abril')
plt.xlabel('Número de Clientes')
plt.ylabel('Tipo de Enlace')

# Agregar etiquetas
for i, v in enumerate(tipo_enlace_count['count']):
    ax.text(v - 25, i, str(v), color='black', ha='left', va='center')

plt.tight_layout()
plt.show()


# In[208]:


for container in ax.containers:
    print(ax.bar_label(container, fmt='%.0f', padding=3))


# ## 5. Relación entre Megas y Tipo de Enlace

# Cruzar megas y tipo de enlace

# In[209]:


# Cruzar megas y tipo de enlace
cross_tab = pd.crosstab(prestador_abril['megas'], prestador_abril['tipo enlace'])


# In[210]:


plt.figure(figsize=(8, 6))
sns.heatmap(cross_tab, annot=True, fmt='d', cmap='YlOrRd')
plt.title('Relación entre Megas y Tipo de Enlace - Abril')
plt.xlabel('Tipo de Enlace')
plt.ylabel('Megas')
plt.tight_layout()
plt.show()


# ## 6. Top 10 de Zonas con Más Clientes

# Top 10 zonas

# In[211]:


# Top 10 zonas
top_zonas = prestador_abril['ZONA'].value_counts().head(10).reset_index()
top_zonas.columns = ['ZONA', 'CLIENTES']


# In[212]:


plt.figure(figsize=(12, 6))
ax = sns.barplot(data=top_zonas, x='CLIENTES', y='ZONA', palette='viridis')
plt.title('Top 10 Zonas con Más Clientes - Abril')
plt.xlabel('Número de Clientes')
plt.ylabel('Zona')

# Agregar etiquetas
for i, v in enumerate(top_zonas['CLIENTES']):
    ax.text(v + 5, i, str(v), color='black', ha='left', va='center')

plt.tight_layout()
plt.show()


# ## 7. Análisis de Compartición (Si es numérica)

# In[325]:


# Si 'comparticion' es numérica
if prestador_abril['comparticion'].dtype in ['int64', 'float64']:
    plt.figure(figsize=(12, 5))

    plt.subplot(1, 2, 1)
    sns.histplot(prestador_abril['comparticion'], bins=20, kde=True)
    plt.title('Distribución de Compartición - Abril')
    plt.xlabel('Compartición')
    plt.ylabel('Frecuencia')

    plt.subplot(1, 2, 2)
    sns.boxplot(y=prestador_abril['comparticion'])
    plt.title('Boxplot de Compartición')
    plt.ylabel('Compartición')

    plt.tight_layout()
    plt.show()

    # Estadísticas descriptivas
    print("Estadísticas descriptivas de compartición:")
    print(prestador_abril['comparticion'].describe())


# ## 8. Mapa de Calor de Correlaciones (Para variables numéricas)

# In[319]:


# Identificar columnas numéricas
numeric_columns = prestador_abril.select_dtypes(include=[np.number]).columns.tolist()

if len(numeric_columns) > 1:
    plt.figure(figsize=(8, 6))
    correlation_matrix = prestador_abril[numeric_columns].corr()
    sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm', center=0)
    plt.title('Mapa de Calor de Correlaciones entre Variables Numéricas - Abril')
    plt.tight_layout()
    plt.show()


# ## 9. Análisis Geográfico Avanzado (Si se tiene coordenadas)

# In[215]:


# Ejemplo si tuvieras coordenadas
# import geopandas as gpd
# from shapely.geometry import Point

# # Crear geometría de puntos
# geometry = [Point(xy) for xy in zip(prestador.longitude, prestador.latitude)]
# geo_df = gpd.GeoDataFrame(prestador, geometry=geometry)

# # Visualizar en mapa
# fig, ax = plt.subplots(figsize=(15, 15))
# geo_df.plot(ax=ax, markersize=5, color='red', marker='o')
# plt.title('Distribución Geográfica de Clientes')
# plt.show()


# ## 10. Análisis de Clientela por Rango de Megas

# In[304]:


bins = [0, 10, 20, 50, 100, 200, 500, 1000]
labels = ['0-10', '11-20', '21-50', '51-100', '101-200', '201-500', '500+']
prestador_abril['rango_megas'] = pd.cut(prestador_abril['megas'], bins=bins, labels=labels, right=False)

# Contar por rango
rango_count = prestador_abril['rango_megas'].value_counts().sort_index().reset_index()
rango_count.columns = ['rango_megas', 'count']

plt.figure(figsize=(8, 4))
ax = sns.barplot(
    data = rango_count,
    x = 'rango_megas',
    y = 'count',
    palette = 'rocket'
)
plt.title('Clientes por Rango de Megas - Abril')
plt.xlabel('Rango de Megas')
plt.ylabel('Número de Clientes')

# Agregar etiquetas
for i, v in enumerate(rango_count['count']):
    ax.text(i, v + 5, str(v), ha='center')

plt.tight_layout()
plt.show()


# In[217]:


pd.cut(prestador_abril['megas'], bins=bins, labels=labels, right=False)


# In[218]:


prestador_abril.head()


# # Lectura de Archivo Mayo

# <b>Mes de evaluación</b>

# In[219]:


nombre_hojas[1]


# In[220]:


prestador_mayo = pd.read_excel('entrada/USUARIOS.xlsx', sheet_name = nombre_hojas[1])


# In[221]:


prestador_mayo.head()


# ## Dimensiones

# In[222]:


prestador_mayo.shape


# In[223]:


prestador_mayo.info()


# In[224]:


prestador_mayo.columns.values


# In[225]:


prestador_mayo.isna().sum()


# In[226]:


print(f'Mes analizado: {nombre_hojas[1]}')


# In[227]:


print (f'Número total de registros de usuarios es: {prestador_mayo.shape[0]}')


# In[228]:


prestador_mayo['tipo enlace'].value_counts().reset_index()


# ## 1. Cantidad Usuarios por Parroquia

# In[229]:


client_parroquia_mayo = (
    prestador_mayo.groupby(['parroquia'])['CLIENTE']
        .count()
        .reset_index()
        .sort_values(by='CLIENTE', ascending=False)
)
client_parroquia_mayo


# In[230]:


client_parroquia_mayo.columns


# <b>Frecuencia de usuarios por parroquia</b>

# In[231]:


frecuencia_user_parroquia_mayo = (
    prestador_mayo.groupby(['parroquia'])['CLIENTE']
        .value_counts()
        .reset_index()
)
frecuencia_user_parroquia_mayo


# In[232]:


frecuencia_user_parroquia_mayo.columns


# <b>Filtro por clientes</b>

# In[233]:


cliente = ['FABIAN RUIZ OLMEDO']
filtro = prestador_mayo['CLIENTE'].isin(cliente)
prestador_mayo[filtro]


# In[234]:


cliente = ['JAVIER LEONARDO CAMPOVERDE RAMIREZ']
filtro = prestador_mayo['CLIENTE'].isin(cliente)
prestador_mayo[filtro]


# In[235]:


cliente = ['JORDAN ALEXANDER MEJIA CRUZ']
filtro = prestador_mayo['CLIENTE'].isin(cliente)
prestador_mayo[filtro]


# In[236]:


cliente = ['JUAN BELTRAN ARBOLEDA']
filtro = prestador_mayo['CLIENTE'].isin(cliente)
prestador_mayo[filtro]


# In[237]:


cliente = ['LENIN CHANATASIG QUIMBITA']
filtro = prestador_mayo['CLIENTE'].isin(cliente)
prestador_mayo[filtro]


# Nota: Realizar comparación entre los clientes activos y estos usuarios

# Realizar comparación de la cantidad de usuarios a lo largo de los 3 meses

# ## 2. Clientes por parroquias

# Datos a graficar

# In[238]:


prestador_count = (
    prestador_mayo.groupby(['parroquia'])['CLIENTE']
        .count()
        .reset_index()
        .sort_values(by='CLIENTE', ascending=False)
)


# In[239]:


# Configuración inicial
plt.figure(figsize=(10, 8))
sns.set_theme(style="whitegrid")

# Barras Horizontal
ax = sns.barplot(
    data=prestador_count,  # DataFrame con los datos agrupados
    y='parroquia',
    x='CLIENTE',
    hue = 'parroquia',
    legend = False,
    orient='h'
)

# Añadir etiquetas (valores de las barras)
for container in ax.containers:
    ax.bar_label(container, fmt='%.0f', padding=3)

# Personalización
plt.title('No. Clientes por Parroquia - Mayo', fontsize=16, pad=20)
plt.xlabel('Número de Clientes', fontsize=12)
plt.ylabel('Parroquia', fontsize=12)

# Ajustar layout y mostrar
plt.tight_layout()
plt.show()



# Etiquetas para las barras

# Etiqueta posición [0]

# In[240]:


ax.bar_label(ax.containers[0], fmt='%.0f', padding=3)  # Agregar valores en las barras


# Etiqueta todas las posiciones

# In[241]:


for container in ax.containers:
    print(ax.bar_label(container, fmt='%.0f', padding=3))


# ## 3. Distribución de Velocidad de Internet (Megas)

# In[242]:


# Análisis de distribución de megas
plt.figure(figsize=(12, 6))
megas_dist = prestador_mayo['megas'].value_counts().sort_index()

# Gráfico Barras Verticales
plt.subplot(1, 2, 1)
sns.barplot(x=megas_dist.index, y=megas_dist.values, palette='coolwarm')
plt.title('Distribución de Planes de Megas - Mayo')
plt.xlabel('Megas')
plt.ylabel('Número de Clientes')
plt.xticks(rotation=45)

# Agregar etiquetas (valores de las barras)
for i, v in enumerate(megas_dist.values):
    plt.text(i, v + 5, str(v), ha='center')

# Gráfico Pie
plt.subplot(1, 2, 2)
plt.pie(megas_dist.values, labels=megas_dist.index, autopct='%1.1f%%')
plt.title('Porcentaje de Clientes por Plan')

# Ajustar Layout y mostrar
plt.tight_layout()
plt.show()


# ## 4. Análisis de Tipo de Enlace

# Distribución por tipo de enlace

# In[243]:


# Distribución por tipo de enlace
tipo_enlace_count = prestador_mayo['tipo enlace'].value_counts().reset_index()
tipo_enlace_count.columns = ['tipo_enlace', 'count']


# In[244]:


plt.figure(figsize=(8, 4))
ax = sns.barplot(data=tipo_enlace_count, y='tipo_enlace', x='count', palette='Set2')
plt.title('Distribución por Tipo de Enlace - Mayo')
plt.xlabel('Número de Clientes')
plt.ylabel('Tipo de Enlace')

# Agregar etiquetas
for i, v in enumerate(tipo_enlace_count['count']):
    ax.text(v - 25, i, str(v), color='black', ha='left', va='center')

plt.tight_layout()
plt.show()


# In[245]:


for container in ax.containers:
    print(ax.bar_label(container, fmt='%.0f', padding=3))


# ## 5. Relación entre Megas y Tipo de Enlace

# Cruzar megas y tipo de enlace

# In[250]:


# Cruzar megas y tipo de enlace
cross_tab = pd.crosstab(prestador_mayo['megas'], prestador_mayo['tipo enlace'])


# In[251]:


plt.figure(figsize=(8, 6))
sns.heatmap(cross_tab, annot=True, fmt='d', cmap='YlOrRd')
plt.title('Relación entre Megas y Tipo de Enlace - Mayo')
plt.xlabel('Tipo de Enlace')
plt.ylabel('Megas')
plt.tight_layout()
plt.show()


# ## 6. Top 10 de Zonas con Más Clientes

# Top 10 zonas

# In[252]:


# Top 10 zonas
top_zonas = prestador_mayo['ZONA'].value_counts().head(10).reset_index()
top_zonas.columns = ['ZONA', 'CLIENTES']


# In[253]:


plt.figure(figsize=(12, 6))
ax = sns.barplot(data=top_zonas, x='CLIENTES', y='ZONA', palette='viridis')
plt.title('Top 10 Zonas con Más Clientes - Mayo')
plt.xlabel('Número de Clientes')
plt.ylabel('Zona')

# Agregar etiquetas
for i, v in enumerate(top_zonas['CLIENTES']):
    ax.text(v + 5, i, str(v), color='black', ha='left', va='center')

plt.tight_layout()
plt.show()


# ## 7. Análisis de Compartición (Si es numérica)

# In[324]:


# Si 'comparticion' es numérica
if prestador_mayo['comparticion'].dtype in ['int64', 'float64']:
    plt.figure(figsize=(12, 5))

    plt.subplot(1, 2, 1)
    sns.histplot(prestador_mayo['comparticion'], bins=20, kde=True)
    plt.title('Distribución de Compartición')
    plt.xlabel('Compartición')
    plt.ylabel('Frecuencia')

    plt.subplot(1, 2, 2)
    sns.boxplot(y=prestador_mayo['comparticion'])
    plt.title('Boxplot de Compartición')
    plt.ylabel('Compartición')

    plt.tight_layout()
    plt.show()

    # Estadísticas descriptivas
    print("Estadísticas descriptivas de compartición:")
    print(prestador_mayo['comparticion'].describe())


# ## 8. Mapa de Calor de Correlaciones (Para variables numéricas)

# In[321]:


# Identificar columnas numéricas
numeric_columns = prestador_mayo.select_dtypes(include=[np.number]).columns.tolist()

if len(numeric_columns) > 1:
    plt.figure(figsize=(8, 6))
    correlation_matrix = prestador_mayo[numeric_columns].corr()
    sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm', center=0)
    plt.title('Mapa de Calor de Correlaciones entre Variables Numéricas - Mayo')
    plt.tight_layout()
    plt.show()


# ## 9. Análisis Geográfico Avanzado (Si se tiene coordenadas)

# In[256]:


# Ejemplo si tuvieras coordenadas
# import geopandas as gpd
# from shapely.geometry import Point

# # Crear geometría de puntos
# geometry = [Point(xy) for xy in zip(prestador.longitude, prestador.latitude)]
# geo_df = gpd.GeoDataFrame(prestador, geometry=geometry)

# # Visualizar en mapa
# fig, ax = plt.subplots(figsize=(15, 15))
# geo_df.plot(ax=ax, markersize=5, color='red', marker='o')
# plt.title('Distribución Geográfica de Clientes')
# plt.show()


# ## 10. Análisis de Clientela por Rango de Megas

# In[259]:


bins = [0, 10, 20, 50, 100, 200, 500, 1000]
labels = ['0-10', '11-20', '21-50', '51-100', '101-200', '201-500', '500+']
prestador_mayo['rango_megas'] = pd.cut(prestador_mayo['megas'], bins=bins, labels=labels, right=False)

# Contar por rango
rango_count = prestador_mayo['rango_megas'].value_counts().sort_index().reset_index()
rango_count.columns = ['rango_megas', 'count']

plt.figure(figsize=(8, 4))
ax = sns.barplot(
    data = rango_count,
    x = 'rango_megas',
    y = 'count',
    palette = 'rocket'
)
plt.title('Clientes por Rango de Megas - Mayo')
plt.xlabel('Rango de Megas')
plt.ylabel('Número de Clientes')

# Agregar etiquetas
for i, v in enumerate(rango_count['count']):
    ax.text(i, v + 5, str(v), ha='center')

plt.tight_layout()
plt.show()


# In[260]:


# Crear categorías de megas
bins = [0, 10, 20, 50, 100, 200, 500, 1000]
labels = ['0-10', '11-20', '21-50', '51-100', '101-200', '201-500', '500+']

pd.cut(prestador_mayo['megas'], bins=bins, labels=labels, right=False)


# In[261]:


prestador_mayo['rango_megas'] = pd.cut(prestador_mayo['megas'], bins=bins, labels=labels, right=False)


# In[262]:


prestador_mayo.head()


# # Lectura de Archivo Junio

# <b>Mes de evaluación</b>

# In[263]:


nombre_hojas[2]


# In[264]:


prestador_junio = pd.read_excel('entrada/USUARIOS.xlsx', sheet_name = nombre_hojas[2])


# In[265]:


prestador_junio.head()


# ## Dimensiones

# In[266]:


prestador_junio.shape


# In[267]:


prestador_junio.info()


# In[268]:


prestador_junio.columns.values


# In[269]:


prestador_junio.isna().sum()


# In[270]:


print(f'Mes analizado: {nombre_hojas[2]}')


# In[271]:


print (f'Número total de registros de usuarios es: {prestador_junio.shape[0]}')


# In[272]:


prestador_junio['tipo enlace'].value_counts().reset_index()


# ## 1. Cantidad Usuarios por Parroquia

# In[316]:


client_parroquia_junio = (
    prestador_junio.groupby(['parroquia'])['CLIENTE']
        .count()
        .reset_index()
        .sort_values(by='CLIENTE', ascending=False)
)
client_parroquia_junio


# In[274]:


client_parroquia_junio.columns


# <b>Frecuencia de usuarios por parroquia</b>

# In[322]:


frecuencia_user_parroquia_junio = (
    prestador_junio.groupby(['parroquia'])['CLIENTE']
        .value_counts()
        .reset_index()
)
frecuencia_user_parroquia_junio


# In[276]:


frecuencia_user_parroquia_junio.columns


# <b>Filtro por clientes</b>

# In[277]:


cliente = ['FABIAN RUIZ OLMEDO']
filtro = prestador_junio['CLIENTE'].isin(cliente)
prestador_junio[filtro]


# In[278]:


cliente = ['JAVIER LEONARDO CAMPOVERDE RAMIREZ']
filtro = prestador_junio['CLIENTE'].isin(cliente)
prestador_junio[filtro]


# In[279]:


cliente = ['JORDAN ALEXANDER MEJIA CRUZ']
filtro = prestador_junio['CLIENTE'].isin(cliente)
prestador_junio[filtro]


# In[280]:


cliente = ['JUAN BELTRAN ARBOLEDA']
filtro = prestador_junio['CLIENTE'].isin(cliente)
prestador_junio[filtro]


# In[281]:


cliente = ['LENIN CHANATASIG QUIMBITA']
filtro = prestador_junio['CLIENTE'].isin(cliente)
prestador_junio[filtro]


# Nota: Realizar comparación entre los clientes activos y estos usuarios

# Realizar comparación de la cantidad de usuarios a lo largo de los 3 meses

# ## 2. Clientes por parroquias

# Datos a graficar

# In[282]:


prestador_count = (
    prestador_junio.groupby(['parroquia'])['CLIENTE']
        .count()
        .reset_index()
        .sort_values(by='CLIENTE', ascending=False)
)


# In[284]:


# Configuración inicial
plt.figure(figsize=(10, 8))
sns.set_theme(style="whitegrid")

# Barras Horizontal
ax = sns.barplot(
    data=prestador_count,  # DataFrame con los datos agrupados
    y='parroquia',
    x='CLIENTE',
    hue = 'parroquia',
    legend = False,
    orient='h'
)

# Añadir etiquetas (valores de las barras)
for container in ax.containers:
    ax.bar_label(container, fmt='%.0f', padding=3)

# Personalización
plt.title('No. Clientes por Parroquia - Junio', fontsize=16, pad=20)
plt.xlabel('Número de Clientes', fontsize=12)
plt.ylabel('Parroquia', fontsize=12)

# Ajustar layout y mostrar
plt.tight_layout()
plt.show()



# Etiquetas para las barras

# Etiqueta posición [0]

# In[285]:


ax.bar_label(ax.containers[0], fmt='%.0f', padding=3)  # Agregar valores en las barras


# Etiqueta todas las posiciones

# In[286]:


for container in ax.containers:
    print(ax.bar_label(container, fmt='%.0f', padding=3))


# ## 3. Distribución de Velocidad de Internet (Megas)

# In[287]:


# Análisis de distribución de megas
plt.figure(figsize=(12, 6))
megas_dist = prestador_junio['megas'].value_counts().sort_index()

# Gráfico Barras Verticales
plt.subplot(1, 2, 1)
sns.barplot(x=megas_dist.index, y=megas_dist.values, palette='coolwarm')
plt.title('Distribución de Planes de Megas - Junio')
plt.xlabel('Megas')
plt.ylabel('Número de Clientes')
plt.xticks(rotation=45)

# Agregar etiquetas (valores de las barras)
for i, v in enumerate(megas_dist.values):
    plt.text(i, v + 5, str(v), ha='center')

# Gráfico Pie
plt.subplot(1, 2, 2)
plt.pie(megas_dist.values, labels=megas_dist.index, autopct='%1.1f%%')
plt.title('Porcentaje de Clientes por Plan')

# Ajustar Layout y mostrar
plt.tight_layout()
plt.show()


# ## 4. Análisis de Tipo de Enlace

# Distribución por tipo de enlace

# In[288]:


# Distribución por tipo de enlace
tipo_enlace_count = prestador_junio['tipo enlace'].value_counts().reset_index()
tipo_enlace_count.columns = ['tipo_enlace', 'count']


# In[289]:


plt.figure(figsize=(8, 4))
ax = sns.barplot(data=tipo_enlace_count, y='tipo_enlace', x='count', palette='Set2')
plt.title('Distribución por Tipo de Enlace - Junio')
plt.xlabel('Número de Clientes')
plt.ylabel('Tipo de Enlace')

# Agregar etiquetas
for i, v in enumerate(tipo_enlace_count['count']):
    ax.text(v - 25, i, str(v), color='black', ha='left', va='center')

plt.tight_layout()
plt.show()


# In[290]:


for container in ax.containers:
    print(ax.bar_label(container, fmt='%.0f', padding=3))


# ## 5. Relación entre Megas y Tipo de Enlace

# Cruzar megas y tipo de enlace

# In[291]:


# Cruzar megas y tipo de enlace
cross_tab = pd.crosstab(prestador_junio['megas'], prestador_junio['tipo enlace'])


# In[292]:


plt.figure(figsize=(8, 6))
sns.heatmap(cross_tab, annot=True, fmt='d', cmap='YlOrRd')
plt.title('Relación entre Megas y Tipo de Enlace - Junio')
plt.xlabel('Tipo de Enlace')
plt.ylabel('Megas')
plt.tight_layout()
plt.show()


# ## 6. Top 10 de Zonas con Más Clientes

# Top 10 zonas

# In[298]:


# Top 10 zonas
top_zonas = prestador_junio['ZONA'].value_counts().head(10).reset_index()
top_zonas.columns = ['ZONA', 'CLIENTES']


# In[299]:


plt.figure(figsize=(12, 6))
ax = sns.barplot(data=top_zonas, x='CLIENTES', y='ZONA', palette='viridis')
plt.title('Top 10 Zonas con Más Clientes - Junio')
plt.xlabel('Número de Clientes')
plt.ylabel('Zona')

# Agregar etiquetas
for i, v in enumerate(top_zonas['CLIENTES']):
    ax.text(v + 5, i, str(v), color='black', ha='left', va='center')

plt.tight_layout()
plt.show()


# ## 7. Análisis de Compartición (Si es numérica)

# In[323]:


# Si 'comparticion' es numérica
if prestador_junio['comparticion'].dtype in ['int64', 'float64']:
    plt.figure(figsize=(12, 5))

    plt.subplot(1, 2, 1)
    sns.histplot(prestador_junio['comparticion'], bins=20, kde=True)
    plt.title('Distribución de Compartición')
    plt.xlabel('Compartición')
    plt.ylabel('Frecuencia')

    plt.subplot(1, 2, 2)
    sns.boxplot(y=prestador_junio['comparticion'])
    plt.title('Boxplot de Compartición')
    plt.ylabel('Compartición')

    plt.tight_layout()
    plt.show()

    # Estadísticas descriptivas
    print("Estadísticas descriptivas de compartición:")
    print(prestador_junio['comparticion'].describe())


# ## 8. Mapa de Calor de Correlaciones (Para variables numéricas)

# In[326]:


# Identificar columnas numéricas
numeric_columns = prestador_junio.select_dtypes(include=[np.number]).columns.tolist()

if len(numeric_columns) > 1:
    plt.figure(figsize=(8, 6))
    correlation_matrix = prestador_junio[numeric_columns].corr()
    sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm', center=0)
    plt.title('Mapa de Calor de Correlaciones entre Variables Numéricas - Junio')
    plt.tight_layout()
    plt.show()


# ## 9. Análisis Geográfico Avanzado (Si se tiene coordenadas)

# In[256]:


# Ejemplo si tuvieras coordenadas
# import geopandas as gpd
# from shapely.geometry import Point

# # Crear geometría de puntos
# geometry = [Point(xy) for xy in zip(prestador.longitude, prestador.latitude)]
# geo_df = gpd.GeoDataFrame(prestador, geometry=geometry)

# # Visualizar en mapa
# fig, ax = plt.subplots(figsize=(15, 15))
# geo_df.plot(ax=ax, markersize=5, color='red', marker='o')
# plt.title('Distribución Geográfica de Clientes')
# plt.show()


# ## 10. Análisis de Clientela por Rango de Megas

# In[327]:


bins = [0, 10, 20, 50, 100, 200, 500, 1000]
labels = ['0-10', '11-20', '21-50', '51-100', '101-200', '201-500', '500+']
prestador_junio['rango_megas'] = pd.cut(prestador_junio['megas'], bins=bins, labels=labels, right=False)

# Contar por rango
rango_count = prestador_junio['rango_megas'].value_counts().sort_index().reset_index()
rango_count.columns = ['rango_megas', 'count']

plt.figure(figsize=(8, 4))
ax = sns.barplot(
    data = rango_count,
    x = 'rango_megas',
    y = 'count',
    palette = 'rocket'
)
plt.title('Clientes por Rango de Megas - Junio')
plt.xlabel('Rango de Megas')
plt.ylabel('Número de Clientes')

# Agregar etiquetas
for i, v in enumerate(rango_count['count']):
    ax.text(i, v + 5, str(v), ha='center')

plt.tight_layout()
plt.show()


# In[328]:


# Crear categorías de megas
bins = [0, 10, 20, 50, 100, 200, 500, 1000]
labels = ['0-10', '11-20', '21-50', '51-100', '101-200', '201-500', '500+']

pd.cut(prestador_junio['megas'], bins=bins, labels=labels, right=False)


# In[329]:


prestador_junio['rango_megas'] = pd.cut(prestador_junio['megas'], bins=bins, labels=labels, right=False)


# In[330]:


prestador_junio.head()


# # Comparaciones entre los Meses

# Cantidad de clientes por mes

# In[313]:


prestador_abril.shape


# In[314]:


prestador_mayo.shape


# In[315]:


prestador_junio.shape


# In[336]:


client_parroquia_abril.dtypes


# In[337]:


type(client_parroquia_abril)


# In[346]:


abril = client_parroquia_abril.copy()
mayo = client_parroquia_mayo.copy()
junio = client_parroquia_junio.copy()


# In[347]:


abril['Mes'] = 'Abril'
mayo['Mes'] = 'Mayo'
junio['Mes'] = 'Junio'


# In[350]:


get_ipython().run_line_magic('who', '')


# In[368]:


# Lista
#clientes = []
#clientes.append([abril, mayo, junio])


# Concatenando el DF

# In[405]:


# DataFrame
clientes = pd.concat([abril, mayo, junio], ignore_index=True)


# In[406]:


clientes.head()


# In[412]:


# Definir el orden de los meses como categorías
meses_orden = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]

# Convertir la columna "Mes" a tipo categórico con orden
clientes["Mes"] = pd.Categorical(
    clientes["Mes"],
    categories=meses_orden,
    ordered=True
)

# Crear tabla pivot (las columnas de meses ahora respetan el orden categórico)
tabla_clientes = clientes.pivot_table(
    values="CLIENTE",
    index="parroquia", # lo que será filas
    columns="Mes", # lo que será columnas
    aggfunc="sum",
    margins=True, # agrega fila y columna "Total"
    margins_name="Total", # nombre que se mostrará en los totales
    observed=True # Muestra solo los meses observados
)#.reset_index()   # bloqueado para que no se generen indices numéricos

# Guardar aparte fila Total
fila_total = tabla_clientes.loc["Total"]

# Eliminar fila Total antes de ordenar
tabla_clientes = tabla_clientes.drop(index="Total")

# Ordenar la tabla
tabla_clientes = tabla_clientes.sort_values(by='Mayo', ascending=False)

# Reagregar la fila Total al final
tabla_clientes.loc["Total"] = fila_total

tabla_clientes


# In[415]:


from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# In[416]:


# === Exportar a Excel ===
ruta_salida = "reporte_clientes.xlsx"
tabla_clientes.to_excel(ruta_salida, sheet_name="Reporte")

# === Dar formato con openpyxl ===
wb = load_workbook(ruta_salida)
ws = wb["Reporte"]

# Encabezados: color de fondo + negrita
fill_header = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = fill_header

# Estilo para Totales (negrita + fondo gris claro)
fill_total = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")

# Formato en fila "Total"
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[0].value == "Total":  # primera celda de la fila
        for cell in row:
            cell.font = Font(bold=True)
            cell.fill = fill_total

# Formato en columna "Total"
for col in range(2, ws.max_column + 1):
    if ws.cell(row=1, column=col).value == "Total":
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = fill_total

# Ajustar ancho de columnas automáticamente
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

# Guardar cambios
wb.save(ruta_salida)

print(f"✅ Reporte exportado con formato en: {ruta_salida}")





